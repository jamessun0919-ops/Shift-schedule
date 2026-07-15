"""
Prototype: 讀取「範例班表.ods」的整月外場班表格狀資料，抽取指定一天的
每人上下班時間，輸出成一個 xlsx 分頁 + 原生橫向長條圖（Gantt 樣式）。

欄位位置目前寫死對應「範例班表.ods」的實際結構，尚未做通用欄位辨識：
- 員工區塊：每人 3 列（列1=早班起訖、列2=晚班起訖、列3=員工編號），
  以 B 欄出現非空字串（姓名）判斷區塊起始列。
- 日期欄：從 D 欄開始，每天固定佔 2 欄（起、訖），第 N 天 = 欄 (4+2*(N-1), 5+2*(N-1))。

用法：python convert_day.py <輸入ods路徑> <日期 1-31> <輸出xlsx路徑>
"""

import sys
import zipfile
import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel

NS = {
    "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
    "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
    "office": "urn:oasis:names:tc:opendocument:xmlns:office:1.0",
}
SHEET_NAME = "整月外場班表"
NAME_COL = 2
FIRST_DAY_COL = 4  # 欄 D
MAX_ROWS_TO_SCAN = 200  # 遠遠超過29~30位員工*3列的範圍，避免掃到後面的空白填充列


def _tag(ns_key, local):
    return f"{{{NS[ns_key]}}}{local}"


def _cell_value(cell_el):
    vtype = cell_el.get(_tag("office", "value-type"))
    if vtype == "float":
        raw = cell_el.get(_tag("office", "value"))
        return float(raw) if raw is not None else None
    if vtype == "string":
        texts = [t.text or "" for t in cell_el.findall(_tag("text", "p"))]
        return "".join(texts) if texts else None
    return None


def read_sheet_rows(ods_path, sheet_name, max_rows):
    z = zipfile.ZipFile(ods_path)
    root = ET.fromstring(z.read("content.xml"))
    table = None
    for t in root.iter(_tag("table", "table")):
        if t.get(_tag("table", "name")) == sheet_name:
            table = t
            break
    if table is None:
        raise ValueError(f"找不到分頁: {sheet_name}")

    rows = []
    for row_el in table.iter(_tag("table", "table-row")):
        if len(rows) >= max_rows:
            break
        cells = {}
        col = 1
        for cell_el in list(row_el):
            repeat = int(cell_el.get(_tag("table", "number-columns-repeated"), "1"))
            if cell_el.tag == _tag("table", "table-cell"):
                value = _cell_value(cell_el)
                if value is not None:
                    for c in range(col, col + repeat):
                        cells[c] = value
            col += repeat
        rows.append(cells)
    return rows


def extract_day_shifts(rows, day):
    start_col = FIRST_DAY_COL + 2 * (day - 1)
    end_col = start_col + 1

    employees = []
    r = 0
    while r < len(rows):
        name = rows[r].get(NAME_COL)
        if isinstance(name, str) and name.strip():
            shift1_row = rows[r]
            shift2_row = rows[r + 1] if r + 1 < len(rows) else {}
            employees.append({
                "name": name.strip(),
                "s1_start": shift1_row.get(start_col),
                "s1_end": shift1_row.get(end_col),
                "s2_start": shift2_row.get(start_col),
                "s2_end": shift2_row.get(end_col),
            })
            r += 3
        else:
            r += 1
    return employees


def build_daily_workbook(employees, day):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{day}"

    headers = ["員工姓名", "早班起", "早班訖", "晚班起", "晚班訖"]
    ws.append(headers)

    time_fmt = "h:mm"
    for emp in employees:
        row = [
            emp["name"],
            _to_time_fraction(emp["s1_start"]),
            _to_time_fraction(emp["s1_end"]),
            _to_time_fraction(emp["s2_start"]),
            _to_time_fraction(emp["s2_end"]),
        ]
        ws.append(row)

    for r in range(2, ws.max_row + 1):
        for c in range(2, 6):
            ws.cell(row=r, column=c).number_format = time_fmt

    # 幫 Gantt 長條圖準備輔助欄：base(起點,隱藏) / dur1(早班長度) / gap(隱藏) / dur2(晚班長度)
    helper_headers = ["base", "dur1", "gap", "dur2"]
    for i, h in enumerate(helper_headers):
        ws.cell(row=1, column=7 + i, value=h)

    day_start = 8 / 24
    for idx, emp in enumerate(employees):
        r = idx + 2
        s1s = _to_time_fraction(emp["s1_start"])
        s1e = _to_time_fraction(emp["s1_end"])
        s2s = _to_time_fraction(emp["s2_start"])
        s2e = _to_time_fraction(emp["s2_end"])

        if s1s is None:
            base, dur1, gap, dur2 = day_start, 0, 0, 0
        else:
            base = s1s
            dur1 = (s1e - s1s) if s1e is not None else 0
            if s2s is not None and s2e is not None:
                gap = s2s - (s1e or s1s)
                dur2 = s2e - s2s
            else:
                gap, dur2 = 0, 0

        ws.cell(row=r, column=7, value=base)
        ws.cell(row=r, column=8, value=dur1)
        ws.cell(row=r, column=9, value=gap)
        ws.cell(row=r, column=10, value=dur2)

    add_gantt_chart(ws, len(employees))
    return wb


def _to_time_fraction(hours):
    if hours is None:
        return None
    return hours / 24


def add_gantt_chart(ws, n_employees):
    chart = BarChart()
    chart.type = "bar"  # 橫向
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = f"{ws.title}日 排班長條圖"
    chart.y_axis.title = None
    chart.x_axis.title = "時間"
    chart.x_axis.number_format = "h:mm"
    chart.x_axis.scaling.min = 8 / 24
    chart.x_axis.scaling.max = 23 / 24

    last_row = n_employees + 1
    cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
    base_ref = Reference(ws, min_col=7, min_row=1, max_row=last_row)
    dur1_ref = Reference(ws, min_col=8, min_row=1, max_row=last_row)
    gap_ref = Reference(ws, min_col=9, min_row=1, max_row=last_row)
    dur2_ref = Reference(ws, min_col=10, min_row=1, max_row=last_row)

    chart.add_data(base_ref, titles_from_data=True)
    chart.add_data(dur1_ref, titles_from_data=True)
    chart.add_data(gap_ref, titles_from_data=True)
    chart.add_data(dur2_ref, titles_from_data=True)
    chart.set_categories(cats)

    # base、gap 兩個輔助數列設為透明（不顯示），只露出 dur1(早班)/dur2(晚班)
    chart.series[0].graphicalProperties.noFill = True
    chart.series[2].graphicalProperties.noFill = True
    chart.series[1].tx = SeriesLabel(v="早班")
    chart.series[3].tx = SeriesLabel(v="晚班")

    chart.height = max(8, n_employees * 0.6)
    chart.width = 24
    ws.add_chart(chart, "L2")


def main():
    if len(sys.argv) != 4:
        print("用法: python convert_day.py <輸入ods路徑> <日期1-31> <輸出xlsx路徑>")
        sys.exit(1)
    ods_path, day_str, out_path = sys.argv[1], sys.argv[2], sys.argv[3]
    day = int(day_str)

    rows = read_sheet_rows(ods_path, SHEET_NAME, MAX_ROWS_TO_SCAN)
    employees = extract_day_shifts(rows, day)
    wb = build_daily_workbook(employees, day)
    wb.save(out_path)
    print(f"完成：{len(employees)} 位員工，輸出至 {out_path}")


if __name__ == "__main__":
    main()
