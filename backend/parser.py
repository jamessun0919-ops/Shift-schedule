import os
import zipfile
import re
import xml.etree.ElementTree as ET
import openpyxl

NS = {
    "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
    "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
    "office": "urn:oasis:names:tc:opendocument:xmlns:office:1.0",
}

def _tag(ns_key, local):
    return f"{{{NS[ns_key]}}}{local}"

def _cell_value(cell_el):
    vtype = cell_el.get(_tag("office", "value-type"))
    if vtype == "float":
        raw = cell_el.get(_tag("office", "value"))
        return float(raw) if raw is not None else None
    if vtype == "date":
        # First try to get text value inside text:p (e.g. "7月1日")
        texts = [t.text or "" for t in cell_el.findall(_tag("text", "p"))]
        val = "".join(texts).strip() if texts else None
        if val:
            return val
        # Fallback to date-value attribute
        return cell_el.get(_tag("office", "date-value"))
    if vtype == "string" or vtype is None:
        texts = [t.text or "" for t in cell_el.findall(_tag("text", "p"))]
        val = "".join(texts).strip() if texts else None
        if val == "":
            return None
        return val
    return None

def parse_time_to_hours(val):
    """
    Converts various time representation formats to float hours (0.0 to 24.0).
    Examples:
        9.0 -> 9.0
        "9.0" -> 9.0
        "09:30" -> 9.5
        "13:00" -> 13.0
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    
    val_str = str(val).strip()
    if not val_str:
        return None
    
    # Try float parsing first (e.g. "9.0" or "9")
    try:
        return float(val_str)
    except ValueError:
        pass
    
    # Try HH:MM or H:MM parsing
    m = re.match(r"^(\d{1,2}):(\d{2})$", val_str)
    if m:
        hours = int(m.group(1))
        minutes = int(m.group(2))
        return hours + minutes / 60.0
        
    return None

def parse_shift_string(val_str):
    """
    Parses strings like "0900-1900(1400-1500)" or "1200-2200" into shift dicts.
    Returns a list of shifts, e.g. [{"start": 9.0, "end": 14.0}, {"start": 15.0, "end": 19.0}].
    """
    if not val_str:
        return []
        
    val_str = str(val_str).strip()
    # Match format like: 0900-1900(1400-1500)
    m = re.match(r"^(\d{4})-(\d{4})(?:\((\d{4})-(\d{4})\))?$", val_str)
    if m:
        s1 = float(m.group(1)[:2]) + float(m.group(1)[2:]) / 60.0
        e1 = float(m.group(2)[:2]) + float(m.group(2)[2:]) / 60.0
        
        if m.group(3) and m.group(4):
            b_s = float(m.group(3)[:2]) + float(m.group(3)[2:]) / 60.0
            b_e = float(m.group(4)[:2]) + float(m.group(4)[2:]) / 60.0
            return [
                {"start": s1, "end": b_s},
                {"start": b_e, "end": e1}
            ]
        else:
            return [{"start": s1, "end": e1}]
            
    # Try simple format like: 09:00-18:00
    m = re.match(r"^(\d{1,2}:\d{2})-(\d{1,2}:\d{2})$", val_str)
    if m:
        s_parts = m.group(1).split(":")
        e_parts = m.group(2).split(":")
        s1 = float(s_parts[0]) + float(s_parts[1]) / 60.0
        e1 = float(e_parts[0]) + float(e_parts[1]) / 60.0
        return [{"start": s1, "end": e1}]
        
    return []

def read_ods_rows(ods_path, sheet_name):
    """
    Safely reads ODS sheet rows without executing macros or external links.
    Returns a list of dicts: {col_idx: value} (1-based col_idx).
    """
    with zipfile.ZipFile(ods_path, "r") as z:
        root = ET.fromstring(z.read("content.xml"))
        
    table = None
    for t in root.iter(_tag("table", "table")):
        if t.get(_tag("table", "name")) == sheet_name:
            table = t
            break
            
    if table is None:
        raise ValueError(f"找不到工作表: {sheet_name}")
        
    rows = []
    for row_el in table.iter(_tag("table", "table-row")):
        row_repeat = int(row_el.get(_tag("table", "number-rows-repeated"), "1"))
        
        cells = {}
        col = 1
        for cell_el in list(row_el):
            col_repeat = int(cell_el.get(_tag("table", "number-columns-repeated"), "1"))
            if cell_el.tag == _tag("table", "table-cell"):
                value = _cell_value(cell_el)
                if value is not None:
                    for c in range(col, col + col_repeat):
                        cells[c] = value
            col += col_repeat
            
        for _ in range(row_repeat):
            rows.append(cells.copy())
            
    return rows

def read_xlsx_rows(xlsx_path, sheet_name):
    """
    Safely reads XLSX sheet rows.
    Uses read_only=True and data_only=True to disable macros/links.
    Returns a list of dicts: {col_idx: value} (1-based col_idx).
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"找不到工作表: {sheet_name}")

        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = {}
            for col_idx, val in enumerate(row, 1):
                if val is not None:
                    cells[col_idx] = val
            rows.append(cells)
        return rows
    finally:
        wb.close()

def get_hidden_columns(file_path, sheet_name):
    """
    Returns a set of 1-based column indices that are hidden in the source file
    (e.g. user collapsed helper columns in Excel/Calc before exporting).
    """
    hidden = set()
    if file_path.endswith(".ods"):
        with zipfile.ZipFile(file_path, "r") as z:
            root = ET.fromstring(z.read("content.xml"))

        table = None
        for t in root.iter(_tag("table", "table")):
            if t.get(_tag("table", "name")) == sheet_name:
                table = t
                break
        if table is None:
            return hidden

        col = 1
        for col_el in table.iter(_tag("table", "table-column")):
            repeat = int(col_el.get(_tag("table", "number-columns-repeated"), "1"))
            visibility = col_el.get(_tag("table", "visibility"), "visible")
            if visibility != "visible":
                for c in range(col, col + repeat):
                    hidden.add(c)
            col += repeat
    else:
        wb = openpyxl.load_workbook(file_path, read_only=False)
        try:
            if sheet_name not in wb.sheetnames:
                return hidden
            ws = wb[sheet_name]
            for dim in ws.column_dimensions.values():
                if dim.hidden and dim.min and dim.max:
                    for c in range(dim.min, dim.max + 1):
                        hidden.add(c)
        finally:
            wb.close()
    return hidden

def extract_top_preview(file_path, sheet_name, max_rows=15):
    """
    Extracts the top rows of a sheet for preview in the secondary feature modal.
    Returns a list of lists of values.
    """
    if file_path.endswith(".ods"):
        rows_data = read_ods_rows(file_path, sheet_name)
    else:
        rows_data = read_xlsx_rows(file_path, sheet_name)
        
    preview = []
    for r_idx in range(min(max_rows, len(rows_data))):
        row = rows_data[r_idx]
        # Convert dictionary to list up to column 40
        max_col = max(row.keys()) if row.keys() else 1
        col_limit = min(40, max_col)
        row_list = [row.get(c, "") for c in range(1, col_limit + 1)]
        preview.append(row_list)
    return preview

def parse_schedule(file_path, template):
    """
    Generic template-driven parser.
    """
    from backend.heuristics import find_block_anchors
    sheet_name = template["sheet_name"]
    if file_path.endswith(".ods"):
        rows = read_ods_rows(file_path, sheet_name)
    else:
        rows = read_xlsx_rows(file_path, sheet_name)
        
    name_col = template["mapping"]["name_col"]
    first_day_col = template["mapping"]["first_day_col"]
    cols_per_day = template["mapping"]["cols_per_day"]
    
    header_row = template.get("header_row_index", 1) # 1-based index
    start_scan_idx = header_row # start checking blocks from header_row (index start_scan_idx - 1 in 0-based)
    
    block_config = template["block"]
    expected_rows = block_config["expected_rows"]
    row_meanings = block_config["row_meanings"]
    
    # 1. Group rows into dynamic employee blocks
    name_row_offset = block_config.get("name_row_offset", 0)
    
    # Groups rows by position within the block grid (phase-anchored on the first
    # genuine name-like cell, not rigid tiling from start_scan_idx -- see
    # backend.heuristics.find_block_anchors), then keeps just the bucket matching
    # the template's known name_row_offset.
    name_row_indices = find_block_anchors(rows, name_col, expected_rows, start_scan_idx, name_row_offset=name_row_offset)

    blocks = []
    for name_row_idx in name_row_indices:
        block_start = name_row_idx - name_row_offset
        block_end = block_start + expected_rows

        block_rows = []
        for r in range(block_start, min(block_end, len(rows))):
            block_rows.append(rows[r])

        name_val = rows[name_row_idx].get(name_col)
        name = str(name_val).strip() if name_val else "未知員工"

        blocks.append({
            "name": name,
            "rows": block_rows
        })
        
    employees = []
    anomalies = []
    
    # 2. Parse each block based on template meanings
    for block in blocks:
        name = block["name"]
        block_rows = block["rows"]
        actual_rows = len(block_rows)
        
        if actual_rows > expected_rows:
            anomalies.append(f"員工 '{name}' 的實際區塊列數 ({actual_rows}) 超出範本預期 ({expected_rows})，多餘列已忽略。")
            block_rows = block_rows[:expected_rows]
        elif actual_rows < expected_rows:
            # Pad with empty dictionaries
            block_rows += [{}] * (expected_rows - actual_rows)
            
        emp_data = {
            "name": name,
            "metadata": {},
            "days": {} # maps 1 to 31
        }
        
        # Count maximum shifts needed
        n_shifts = 0
        for m in row_meanings:
            if m["type"] == "shift":
                n_shifts = max(n_shifts, m["index"] + 1)
            elif m["type"] in ["shift_start", "shift_end"]:
                n_shifts = max(n_shifts, m["index"] + 1)
            elif m["type"] == "shift_string":
                n_shifts = max(n_shifts, 2) # shift string parsing can output up to 2 shifts
                
        # Initialize days 1-31 with empty shifts list
        for d in range(1, 32):
            emp_data["days"][d] = {"shifts": [{"start": None, "end": None} for _ in range(n_shifts)]}
            
        for row_idx, meaning in enumerate(row_meanings):
            row_data = block_rows[row_idx]
            
            if meaning["type"] == "shift":
                shift_idx = meaning["index"]
                for d in range(1, 32):
                    start_col = first_day_col + (d - 1) * cols_per_day
                    end_col = start_col + 1
                    s_val = parse_time_to_hours(row_data.get(start_col))
                    e_val = parse_time_to_hours(row_data.get(end_col))
                    if shift_idx < len(emp_data["days"][d]["shifts"]):
                        emp_data["days"][d]["shifts"][shift_idx] = {"start": s_val, "end": e_val}
                    
            elif meaning["type"] == "shift_start":
                shift_idx = meaning["index"]
                for d in range(1, 32):
                    col = first_day_col + (d - 1) * cols_per_day
                    s_val = parse_time_to_hours(row_data.get(col))
                    if shift_idx < len(emp_data["days"][d]["shifts"]):
                        emp_data["days"][d]["shifts"][shift_idx]["start"] = s_val
                    
            elif meaning["type"] == "shift_end":
                shift_idx = meaning["index"]
                for d in range(1, 32):
                    col = first_day_col + (d - 1) * cols_per_day
                    e_val = parse_time_to_hours(row_data.get(col))
                    if shift_idx < len(emp_data["days"][d]["shifts"]):
                        emp_data["days"][d]["shifts"][shift_idx]["end"] = e_val
                    
            elif meaning["type"] == "shift_string":
                for d in range(1, 32):
                    col = first_day_col + (d - 1) * cols_per_day
                    val_str = row_data.get(col)
                    parsed_shifts = parse_shift_string(val_str)
                    for s_idx, s in enumerate(parsed_shifts):
                        if s_idx < len(emp_data["days"][d]["shifts"]):
                            emp_data["days"][d]["shifts"][s_idx] = s
                            
            elif meaning["type"] == "metadata":
                meta_name = meaning["name"]
                meta_col = meaning.get("col", name_col)
                emp_data["metadata"][meta_name] = row_data.get(meta_col)
                
        employees.append(emp_data)
        
    # 3. Health Check
    is_healthy, health_msg = perform_health_check(employees)
    if not is_healthy:
        anomalies.append(f"健康檢查警報: {health_msg}")
        
    return {
        "employees": employees,
        "anomalies": anomalies,
        "is_healthy": is_healthy
    }

def perform_health_check(employees):
    """
    Validates if the parsed data makes sense.
    """
    if not employees:
        return False, "未解析出任何員工資料。"
        
    total_shifts_checked = 0
    invalid_shifts = 0
    
    for emp in employees:
        for d, day_info in emp["days"].items():
            for shift in day_info["shifts"]:
                s = shift["start"]
                e = shift["end"]
                
                if s is not None or e is not None:
                    total_shifts_checked += 1
                    
                    # A valid shift time must be between 0.0 and 24.0
                    # and end must be >= start
                    if s is not None and (s < 0.0 or s > 24.0):
                        invalid_shifts += 1
                    elif e is not None and (e < 0.0 or e > 24.0):
                        invalid_shifts += 1
                    elif s is not None and e is not None and e < s:
                        invalid_shifts += 1
                        
    if total_shifts_checked == 0:
        return False, "班表中未包含任何排班上下班時間數據。"
        
    invalid_ratio = invalid_shifts / total_shifts_checked
    if invalid_ratio > 0.3: # More than 30% of shifts are invalid
        return False, f"有過高比例 ({invalid_ratio:.1%}) 的時間數據無法正確解析，請確認範本欄位對照是否正確。"
        
    return True, "健康檢查通過。"
