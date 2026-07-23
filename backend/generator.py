import os
import re
import shutil
import zipfile
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.label import DataLabel, DataLabelList
from openpyxl.styles import PatternFill, Font, Alignment

# 班別長條的統一顏色，對應網頁預覽的 --series-1（前後段班別性質上無差異，不用顏色區分）
DUR_SERIES_COLOR = "2A78D6"

def compute_hourly_headcount(employees, day, axis_start, axis_end):
    """
    對 [axis_start, axis_end) 內每個整點小時 h，統計「班別與 [h, h+1) 區間重疊」
    的員工人數，以重疊時數佔該小時的比例計算（例如 9:30 上班，9 點這格記為
    0.5，非整點上下班時為小數）。同一員工在同一小時內的多段班別重疊時數會
    加總。回傳依小時排序的人數列表，長度為 axis_end - axis_start。
    """
    counts = []
    for h in range(axis_start, axis_end):
        count = 0.0
        for emp in employees:
            day_info = emp["days"].get(day)
            if not day_info:
                continue
            for shift in day_info["shifts"]:
                s, e = shift.get("start"), shift.get("end")
                if s is None or e is None or e <= s:
                    continue
                overlap = min(e, h + 1) - max(s, h)
                if overlap > 0:
                    count += overlap
        counts.append(count)
    return counts

_DLBL_NUMFMT_RE = re.compile(rb'(<numFmt formatCode="&quot;[^"]*&quot;")\s*/>')

def _fix_datalabel_numfmt_sourcelinked(xlsx_path):
    """
    openpyxl 的 DataLabel.numFmt 只能設定 formatCode，沒有介面可設定 OOXML
    規格要求的 sourceLinked 屬性；沒有明確標示 sourceLinked="0"，Excel／
    LibreOffice 會判定「沿用來源儲存格格式」而忽略我們自訂的格式碼字面值
    （用來顯示每小時人數，見 add_xlsx_gantt_chart）。存檔後直接補這個屬性
    進圖表 XML（openpyxl 沒有 API 可用，只能後處理）。
    """
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        items = {info.filename: zin.read(info.filename) for info in zin.infolist()}
        infos = zin.infolist()

    for name in items:
        if name.startswith("xl/charts/chart") and name.endswith(".xml"):
            items[name] = _DLBL_NUMFMT_RE.sub(rb'\1 sourceLinked="0" />', items[name])

    with zipfile.ZipFile(xlsx_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for info in infos:
            zout.writestr(info, items[info.filename])

def _get_n_shifts(row_meanings):
    n_shifts = 0
    for m in row_meanings:
        if m["type"] == "shift":
            n_shifts = max(n_shifts, m.get("index", 0) + 1)
        elif m["type"] in ["shift_start", "shift_end"]:
            n_shifts = max(n_shifts, m.get("index", 0) + 1)
        elif m["type"] == "shift_string":
            n_shifts = max(n_shifts, 2)
    return max(1, n_shifts)

def generate_xlsx(employees, output_path, template_config, month=None):
    """
    Generates a monthly XLSX workbook. Each day is a sheet containing the schedule table and Gantt chart.
    """
    display_config = template_config.get("display") or {}
    axis_start = display_config.get("axis_start", 8)
    axis_end = display_config.get("axis_end", 24)

    wb = Workbook()
    wb.remove(wb.active) # Remove default sheet

    for day in range(1, 32):
        ws = wb.create_sheet(title=f"{day}")
        
        # 1. Headers
        headers = ["員工姓名"]
        block_config = template_config["block"]
        row_meanings = block_config["row_meanings"]
        
        # Count shifts
        n_shifts = _get_n_shifts(row_meanings)
        for s_idx in range(1, n_shifts + 1):
            headers.extend([f"班別{s_idx}起", f"班別{s_idx}訖"])
            
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # 2. Populate employee raw shifts
        # 第 2 列保留給「上班人數」虛擬列（見下方甘特圖姓名軸新增列），真實
        # 員工資料從第 3 列開始，姓名軸標籤/座標軸相關列位須同步位移。
        ws.cell(row=2, column=1, value="上班人數")
        time_fmt = "h:mm"
        for idx, emp in enumerate(employees):
            r = idx + 3
            row_data = [emp["name"]]

            day_info = emp["days"][day]
            for s_idx in range(n_shifts):
                shift = day_info["shifts"][s_idx] if s_idx < len(day_info["shifts"]) else {"start": None, "end": None}

                start_frac = (shift["start"] / 24) if shift["start"] is not None else None
                end_frac = (shift["end"] / 24) if shift["end"] is not None else None
                row_data.extend([start_frac, end_frac])

            ws.append(row_data)

            # Format times
            for c_idx in range(2, 2 + 2 * n_shifts):
                cell = ws.cell(row=r, column=c_idx)
                cell.number_format = time_fmt
                cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.cell(row=r, column=1).font = Font(name="Microsoft JhengHei", size=10, bold=True)

        ws.cell(row=2, column=1).font = Font(name="Microsoft JhengHei", size=10, bold=True)

        # 3. Build helper columns for Gantt chart
        start_helper_col = 2 + 2 * n_shifts
        helper_headers = ["base"]
        for s_idx in range(1, n_shifts + 1):
            helper_headers.append(f"dur{s_idx}")
            if s_idx < n_shifts:
                helper_headers.append(f"gap{s_idx}")

        # 每小時人數輔助欄：每個整點一欄，固定長度 1 小時（1/24），確保與時間軸
        # 刻度對齊；真實員工列全部填 0（見下方），只有「上班人數」列（第2列）
        # 會依序填滿整個座標軸範圍。
        hourly_hours = list(range(axis_start, axis_end))
        for h in hourly_hours:
            helper_headers.append(f"cnt{h}")

        for h_idx, h_name in enumerate(helper_headers):
            ws.cell(row=1, column=start_helper_col + h_idx, value=h_name)

        # 輔助欄（base/dur/gap/cnt）僅供甘特圖繪製使用，對使用者無意義，隱藏但保留資料。
        for h_idx in range(len(helper_headers)):
            ws.column_dimensions[_col_letter(start_helper_col + h_idx)].hidden = True

        day_start = axis_start / 24
        cnt_start_col = start_helper_col + 2 * n_shifts

        # 「上班人數」虛擬列（第2列）：真實班別欄位全部填 0（不畫班別長條），
        # 每小時欄依序填 1/24（固定 1 小時長度），從 axis_start 開始接續填滿到
        # axis_end，長條本身透明、只靠資料標籤顯示人數（見 add_xlsx_gantt_chart）。
        ws.cell(row=2, column=start_helper_col, value=day_start)
        for s_idx in range(2 * n_shifts - 1):
            ws.cell(row=2, column=start_helper_col + 1 + s_idx, value=0.0)
        for h_idx in range(len(hourly_hours)):
            ws.cell(row=2, column=cnt_start_col + h_idx, value=1 / 24)

        for idx, emp in enumerate(employees):
            r = idx + 3
            day_info = emp["days"][day]

            active_shifts = []
            for s in day_info["shifts"]:
                if s["start"] is not None and s["end"] is not None:
                    active_shifts.append(s)

            active_shifts.sort(key=lambda x: x["start"])

            segments = []
            if not active_shifts:
                base = day_start
                durations_and_gaps = [0.0] * (2 * n_shifts - 1)
            else:
                base = active_shifts[0]["start"] / 24
                durations_and_gaps = [(active_shifts[0]["end"] - active_shifts[0]["start"]) / 24]

                for s_idx in range(1, len(active_shifts)):
                    gap = (active_shifts[s_idx]["start"] - active_shifts[s_idx - 1]["end"]) / 24
                    dur = (active_shifts[s_idx]["end"] - active_shifts[s_idx]["start"]) / 24
                    durations_and_gaps.extend([gap, dur])

                needed_segments = 2 * n_shifts - 1
                if len(durations_and_gaps) < needed_segments:
                    durations_and_gaps.extend([0.0] * (needed_segments - len(durations_and_gaps)))

            ws.cell(row=r, column=start_helper_col, value=base)
            for s_idx, val in enumerate(durations_and_gaps):
                ws.cell(row=r, column=start_helper_col + 1 + s_idx, value=val)
            # 真實員工不參與「上班人數」列的每小時欄，全部填 0。
            for h_idx in range(len(hourly_hours)):
                ws.cell(row=r, column=cnt_start_col + h_idx, value=0.0)

        hourly_counts = compute_hourly_headcount(employees, day, axis_start, axis_end)
        add_xlsx_gantt_chart(ws, len(employees), n_shifts, start_helper_col, axis_start, axis_end, month, hourly_counts)

    wb.save(output_path)
    _fix_datalabel_numfmt_sourcelinked(output_path)
    print(f"Generated XLSX schedule: {output_path}")

def add_xlsx_gantt_chart(ws, n_employees, n_shifts, start_helper_col, axis_start=8, axis_end=24, month=None, hourly_counts=None):
    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "stacked"
    chart.overlap = 100
    month_prefix = f"{month}月" if month else ""
    chart.title = f"{month_prefix}{ws.title}日 排班長條圖"
    # 輔助欄（base/dur/gap/cnt）在上面被隱藏，但圖表仍需讀取其資料繪圖：
    # Excel 圖表預設 plotVisOnly=true（只畫可見儲存格），必須關閉。
    chart.visible_cells_only = False
    chart.legend = None

    hourly_counts = hourly_counts or []
    n_hour_segs = len(hourly_counts)
    # 姓名軸多一列「上班人數」虛擬列（見 generate_xlsx），所有列數計算加 1。
    last_row = n_employees + 2

    total_series = 2 * n_shifts
    total_series_all = total_series + n_hour_segs
    for s_idx in range(total_series_all):
        col = start_helper_col + s_idx
        ref = Reference(ws, min_col=col, min_row=1, max_row=last_row)
        chart.add_data(ref, titles_from_data=True)

    # set_categories() 必須在 add_data() 之後呼叫：呼叫當下 chart.series 是空的，
    # 姓名分類永遠不會被寫進任何 series，長條圖左側會變成 Excel 自動編號 1..N。
    cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
    chart.set_categories(cats)

    # 姓名軸（x_axis，類別軸）：預設由下往上排列，reverse 成由上往下。
    chart.x_axis.title = None
    chart.x_axis.scaling.orientation = "maxMin"
    # 類別軸標籤間隔預設是 Excel/LibreOffice 的「自動」，員工數一多、圖表高度
    # 顯得不夠時，渲染軟體會自動跳過部分姓名標籤（並非資料遺漏，只是沒顯示）。
    # 強制間隔為 1，確保每一位員工的姓名都會顯示。
    chart.x_axis.tickLblSkip = 1
    chart.x_axis.tickMarkSkip = 1

    # 實測發現 LibreOffice 即使 tickLblSkip=1 仍會依「圖表高度 ÷ 類別數」自行
    # 算間隔並跳過標籤，需要圖表本身夠高才會真的顯示全部姓名。以 10 個類別的
    # 高度為下限（避免人少時圖表過矮），不設上限、依類別數（員工+人數列）等比放大。
    chart.height = max(10, n_employees + 1) * 0.8

    # 時間軸（y_axis，數值軸）：原本誤設定在 x_axis 上，導致真正的數值軸完全沒有
    # 格式化（顯示 0~1 的原始小數）；改回設在 y_axis，刻度間隔比照網頁版每小時一格。
    # 用 "[h]:mm"（經過時間格式）而非 "h:mm"：後者在數值剛好等於 1.0（=24:00）
    # 時，Excel 會把它當成隔天 0:00 顯示，導致座標軸標籤回繞成 "0:00"。
    chart.y_axis.title = "時間"
    chart.y_axis.number_format = "[h]:mm"
    chart.y_axis.scaling.min = axis_start / 24
    chart.y_axis.scaling.max = axis_end / 24
    chart.y_axis.majorUnit = 1 / 24

    chart.series[0].graphicalProperties.noFill = True

    # 前後段班別性質上無差異，故所有班別長條統一同一色（與網頁預覽 --series-1 一致），
    # 不用 Excel 預設的逐 series 自動配色。
    for s_idx in range(1, total_series):
        is_dur = (s_idx % 2 == 1)
        if is_dur:
            shift_num = (s_idx // 2) + 1
            chart.series[s_idx].tx = SeriesLabel(v=f"班別 {shift_num}")
            chart.series[s_idx].graphicalProperties.solidFill = DUR_SERIES_COLOR
        else:
            chart.series[s_idx].graphicalProperties.noFill = True

    # 每小時人數（姓名軸最上方「上班人數」虛擬列）：區段本身透明，只靠資料
    # 標籤顯示人數。區段實際數值固定為 1 小時長度（用來對齊時間軸刻度，不能
    # 直接當人數顯示），改用 DataLabel.numFmt 的引號字面值技巧（Excel 自訂數字
    # 格式支援用雙引號包住的文字固定顯示，不受實際數值影響）覆蓋顯示文字；
    # 「上班人數」列是姓名軸第一個類別（idx=0），只對該點加標籤。
    # 實測發現：不明確關閉 showCatName/showSerName，LibreOffice 會把類別名稱／
    # 數列名稱一起串進標籤文字，變成「上班人數;cnt8;0.0417」的疊字亂碼，須明確關閉。
    # 且這個關閉必須設在 dLbls 這一層（適用其他未被 idx=0 覆蓋的點，即真實員工
    # 那些人數欄=0的點），否則其他點會落回「未設定=顯示」的預設值，在真實
    # 員工列的長條上疊出「Alice;cnt10;0」這類多餘文字。
    dlbls_common = dict(
        showCatName=False, showSerName=False,
        showLegendKey=False, showPercent=False, showBubbleSize=False,
    )
    for h_idx in range(n_hour_segs):
        s_idx = total_series + h_idx
        series = chart.series[s_idx]
        series.graphicalProperties.noFill = True
        count_val = round(hourly_counts[h_idx], 2)
        count_label = str(int(count_val)) if count_val == int(count_val) else f"{count_val:g}"
        series.dLbls = DataLabelList(
            dLbl=[DataLabel(idx=0, numFmt=f'"{count_label}"', showVal=True, **dlbls_common)],
            showVal=False, **dlbls_common,
        )

    chart.width = 24

    # 上方時間軸（次座標軸）：使用者要求時間軸比照網頁預覽顯示在圖表上方。
    # 實測發現直接調整同一數值軸的 crosses/axPos/tickLblPos 這幾個 OOXML
    # 屬性，在 LibreOffice、Google 試算表皆會被忽略、軸線位置不會改變，故改用
    # 「新增一組次座標數值軸」這個較多軟體都支援的做法：另建一個隱藏 series 的
    # 圖表，設定與主座標相同的時間刻度，用 openpyxl 的 chart += chart2 合併，
    # 使其成為次座標軸；經 headless 渲染確認會顯示在圖表上方，且姓名軸（類別軸）
    # 因兩個圖表共用同一組類別，不會被重複畫出。使用者確認上下兩軸都保留。
    top_axis_chart = BarChart()
    top_axis_chart.type = "bar"
    top_ref = Reference(ws, min_col=start_helper_col, min_row=1, max_row=last_row)
    top_axis_chart.add_data(top_ref, titles_from_data=True)
    top_axis_chart.set_categories(cats)
    top_axis_chart.series[0].graphicalProperties.noFill = True
    top_axis_chart.y_axis.axId = 200
    top_axis_chart.y_axis.title = None
    top_axis_chart.y_axis.number_format = "[h]:mm"
    top_axis_chart.y_axis.scaling.min = axis_start / 24
    top_axis_chart.y_axis.scaling.max = axis_end / 24
    top_axis_chart.y_axis.majorUnit = 1 / 24
    top_axis_chart.y_axis.crosses = "max"
    top_axis_chart.x_axis.delete = True
    chart += top_axis_chart

    # 用 _col_letter() 而非 chr(65+n)：新增每小時人數輔助欄後，欄位總數容易超過
    # 26 欄（單一英文字母上限），chr() 換算會產生非法欄位代號。
    chart_col = _col_letter(start_helper_col + total_series_all + 1)
    ws.add_chart(chart, f"{chart_col}2")


def _col_letter(col_idx):
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result

